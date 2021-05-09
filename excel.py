
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import time
import json
from datetime import datetime

print("Inputting data to excel file...")

PRINT_ALL = False

def days_between(d1, d2):
    d1 = datetime.strptime(d1, "%Y-%m-%d")
    d2 = datetime.strptime(d2, "%Y-%m-%d")
    return abs((d2 - d1).days)

CG = []
CMC = []

with open('CG.txt', 'r') as f:
    CG = json.loads(f.read())
with open('CMC.txt', 'r') as f:
    CMC = json.loads(f.read())

book = Workbook()
sheet = book.active
sheet.title = "CG and CMC"

sheet['A1'] = "Name"
sheet['B1'] = "CG Time"
sheet['C1'] = "CG Date"
sheet['D1'] = "CMC Time"
sheet['E1'] = "CMC Date"
sheet['F1'] = "Diffrence"

sheet['A1'].fill = PatternFill(start_color="55C7CE", fill_type = "solid")
sheet['B1'].fill = PatternFill(start_color="BBFF99", fill_type = "solid")
sheet['C1'].fill = PatternFill(start_color="BBFFDD", fill_type = "solid")
sheet['D1'].fill = PatternFill(start_color="AAC7CC", fill_type = "solid")
sheet['E1'].fill = PatternFill(start_color="AAC7FF", fill_type = "solid")
sheet['F1'].fill = PatternFill(start_color="FFCCBB", fill_type = "solid")

sheet.column_dimensions['A'].width = 25
sheet.column_dimensions['B'].width = 10
sheet.column_dimensions['C'].width = 12
sheet.column_dimensions['D'].width = 10
sheet.column_dimensions['E'].width = 12
sheet.column_dimensions['F'].width = 12

row_num = 0
if(PRINT_ALL):
    for i in range(len(CG)):
        name = CG[i]["name"]
        sheet.cell(row=i+2, column=1).value = name
        sheet.cell(row=i+2, column=2).value = CG[i]["time"]
        sheet.cell(row=i+2, column=3).value = CG[i]["date"]
        for c in CMC:
            if(c["name"] == name):
                sheet.cell(row=i+2, column=4).value = c["time"]
                sheet.cell(row=i+2, column=5).value = c["date"]
                sheet.cell(row=i+2, column=6).value = days_between(c["date"], CG[i]["date"])
else:
    for i in range(len(CG)):
        name = CG[i]["name"]  
        for c in CMC:
            if(c["name"] == name):
                sheet.cell(row=row_num+2, column=1).value = name
                sheet.cell(row=row_num+2, column=2).value = CG[i]["time"]
                sheet.cell(row=row_num+2, column=3).value = CG[i]["date"]
                sheet.cell(row=row_num+2, column=4).value = c["time"]
                sheet.cell(row=row_num+2, column=5).value = c["date"]
                sheet.cell(row=row_num+2, column=6).value = days_between(c["date"], CG[i]["date"])
                row_num += 1

book.save("sample.xlsx")

print("Done.")





