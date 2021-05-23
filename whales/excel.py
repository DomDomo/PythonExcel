from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension
from string import ascii_uppercase

border_style = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

grey_background = PatternFill(start_color="D8D8D9", fill_type = "solid")

# Adds commas to large numbers: 123456789 -> 123,456,789
def add_commas(num):
    return '{:,}'.format(num)

def readable_number(num):
    rounded_integer = round(int(num.replace(",", "")), -12) 
    trillion = 1000000000000
    with_word = str(round(rounded_integer, -12) / trillion).split(".")[0] + " Trillion"
    return with_word

def find_difference(holder):
    d1 = int(holder["quantity"].replace(",", ""))
    d2 = int(holder["old_quantity"].replace(",", ""))
    return d1-d2

def list_holders(sheet, row, column, holders):
    rank = 1
    coin_sum = 0
    difference_sum = 0
    percent_sum = 0
    for holder in holders:
        for i in range(1, column+1):
            sheet.cell(row=row, column=i).border = border_style

        # Rank
        sheet.cell(row=row, column=1).value = rank

        # Top Wallets
        sheet.cell(row=row, column=2).value = holder["address"]

        # Amount (in coins)
        sheet.cell(row=row, column=3).value = holder["quantity"]
        sheet.cell(row=row, column=3).alignment = Alignment(horizontal='right')
        coin_sum += int(holder["quantity"].replace(",", ""))

        # Readable Amount
        sheet.cell(row=row, column=4).value = readable_number(holder["quantity"])
        sheet.cell(row=row, column=4).alignment = Alignment(horizontal='right')

        # % of Supply
        sheet.cell(row=row, column=5).value = holder["percentage"]
        sheet.cell(row=row, column=5).alignment = Alignment(horizontal='right')
        percent_sum += float(holder["percentage"][:-2])

        trillion = 1000000000000
        difference =  find_difference(holder);
        formatted_difference = add_commas(abs(difference))
        
        if abs(difference) >= trillion:
            difference_sum += difference
            
            # More/Less
            more_or_less = " More" if difference > 0 else " Less"
            sheet.cell(row=row, column=6).value = readable_number(formatted_difference) + more_or_less

            # 24HT Transaction Amounts
            sheet.cell(row=row, column=7).value = formatted_difference
            sheet.cell(row=row, column=7).alignment = Alignment(horizontal='right')

            # Bold Losses
            if "Less" in more_or_less:
                sheet.cell(row=row, column=6).font = Font(bold=True)
                sheet.cell(row=row, column=7).font = Font(bold=True)

            # Color grey
            for i in range(1, column+1):
                sheet.cell(row=row, column=i).fill = grey_background

        row += 1
        rank += 1

    return [coin_sum, difference_sum, percent_sum]

def save_to_excel(holders, old_whales):

    book = Workbook()
    sheet = book.active
    sheet.title = "Holders"

    names = [
        [5, ""],
        [50, "Top Wallets"],
        [25, "Amount (in coins)"],
        [25, "Readable Amount"],
        [15, "% of Supply"],
        [22, "More/Less"],
        [25, "24HT Transaction Amounts"]
    ]

    column = len(names);

    # Column names
    name_i = 0
    for c in ascii_uppercase[:column]:
        xy = c + "1"
        sheet.column_dimensions[c].width = names[name_i][0]
        sheet[xy] = names[name_i][1]
        sheet[xy].alignment = Alignment(horizontal='center')
        sheet[xy].border = border_style
        name_i += 1


    # Start holder info from the second row
    row = 2

    coin_sum, difference_sum, percent_sum = list_holders(sheet, row, column, holders)

    # Totals for top 50-------------------------------------------------------
    total_row = len(holders)+3
    for i in range(2, column+1):
            sheet.cell(row=total_row, column=i).border = border_style

    sheet.cell(row=total_row, column=2).value = "Total's for top 50 (actually 48)"
    sheet.cell(row=total_row, column=2).alignment = Alignment(horizontal='right')

    formatted_sum = add_commas(coin_sum)
    sheet.cell(row=total_row, column=3).value = formatted_sum
    sheet.cell(row=total_row, column=3).alignment = Alignment(horizontal='right')

    sheet.cell(row=total_row, column=4).value = readable_number(formatted_sum)
    sheet.cell(row=total_row, column=4).alignment = Alignment(horizontal='right')

    sheet.cell(row=total_row, column=5).value = str(round(percent_sum, 2)) + "%"
    sheet.cell(row=total_row, column=5).alignment = Alignment(horizontal='right')

    formatted_difference_sum = add_commas(difference_sum)

    sheet.cell(row=total_row, column=6).value = readable_number(formatted_difference_sum)

    sheet.cell(row=total_row, column=7).value = formatted_difference_sum

    # Old whales--------------------------------------------------------------
    row = total_row+2
    sheet.cell(row=row, column=2).value = "Old Whales"
    sheet.cell(row=row, column=2).alignment = Alignment(horizontal='center')

    for i in range(1, column+1):
        sheet.cell(row=row, column=i).fill = PatternFill(start_color="ABABAB", fill_type = "solid")
        sheet.cell(row=row, column=i).border = border_style


    left_coin_sum, left_diffrence_sum, left_percent_sum = list_holders(sheet, row+1, column, old_whales)

    coin_sum += left_coin_sum
    difference_sum += left_diffrence_sum
    percent_sum += left_percent_sum

    total_row = row + len(old_whales) + 2
    for i in range(4, column+1):
            sheet.cell(row=total_row, column=i).border = border_style
            sheet.cell(row=total_row, column=i).fill = grey_background
            sheet.cell(row=total_row, column=i).font = Font(bold=True)

    sheet.cell(row=total_row, column=5).value = "Total loss / gain from all whales"
    sheet.cell(row=total_row, column=5).alignment = Alignment(horizontal='right')

    formatted_difference_sum = add_commas(difference_sum)

    sheet.cell(row=total_row, column=6).value = readable_number(formatted_difference_sum)

    sheet.cell(row=total_row, column=7).value = formatted_difference_sum

    book.save("holders.xlsx")




