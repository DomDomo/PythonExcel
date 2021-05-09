
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import time
import json
from datetime import datetime, timedelta


CG = []
CMC = []

with open('CG.txt', 'r') as f:
    CG = json.loads(f.read())
with open('CMC.txt', 'r') as f:
    CMC = json.loads(f.read())

book = Workbook()
ws = book.active
ws.title = "Test"

a = (datetime.date(datetime.now()) + timedelta(days=2)).strftime("%m-%d-%Y")
b = (datetime.date(datetime.now()) + timedelta(hours=2)).strftime("%m-%d-%Y")
c = (datetime.date(datetime.now()) + timedelta(hours=26)).strftime("%m-%d-%Y")
print(a)
ws['A1'] = a
ws['A2'] = b
ws['A3'] = c

book.save("data.xlsx")






